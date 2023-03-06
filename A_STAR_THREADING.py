
from numpy.lib.arraysetops import setdiff1d

import pandas as pd
import numpy
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import glob
import numpy
import re
import report2pygame
import Report1_thread
import report1pygame
import datadump
import report1datadump
import input
import sys
from collections import Counter
import copy
import os
import shutil
import threading
import concurrent.futures
import multiprocessing
#from multiprocessing import Pool
from concurrent.futures import ProcessPoolExecutor as Pool

start = time.time()

# Layout Paths
ROOT = os.getcwd()
EXCEL_PATH = os.path.join(ROOT, "LTA_2-1Scale.xlsx")  # LAYOUT PATH
REPORT_PATH = os.path.join(ROOT, "sample.xlsx")  # INPUT PATH
COMP_PATH=os.path.join(ROOT, "Input", "VPP_COMP Data.csv") #VPP COMP PATH
SAMPLEFUNCTION_PATH=os.path.join(ROOT, "Input", "samplefunction.csv") #SAMPLE FUNCTION
PARTLIST_PATH=os.path.join(ROOT, "Input", "PartsList.csv")

# INPUTS
WORKCENTRE = ["16WSLT39"]  #GIVE INPUT
SOURCE_TEXT_ARRAY = ["D9"] #models required
MECH_CLASSES="ALL"
#MECH_CLASSES = ["LTAF40", "LTAL42", "LTAR41"]
REPORT1 = 1
REPORT2 = 0
CORES = multiprocessing.cpu_count()


# HIGHRISE LINE 
#START_CELL = "Ahj109"
#END_CELL = "AVH308"
#IMAGE_PATH=os.path.join(ROOT,"LARGElayout.jpg")  #Image of the layout if LARGELINE
#SHEET_NAME="Grid-Loc-HighRise"


# EMR LINE
START_CELL = "KA55"
END_CELL = "AFZ308"
# Image of the layout if EMR LINE
IMAGE_PATH = os.path.join(ROOT, "EMRLayout.png")
#SHEET_NAME = "Grid-Loc-EMR"
SHEET_NAME = "Grid-Loc-Fender"
#SHEET_NAME = "Grid-Loc-Build"

# pygame global variable inputs (Visualization)
SCALE = 0.5             # 1 grid in excel= (0.5 ft * 0.5 ft)
GAP = 1  # (1ft)
CHECK_CELL = int(GAP/SCALE)  # number of cells to search
STRIDE = 4  # 4 inches per second
STRIDE_MIN = 240  # inches per minute
PATH_COLOR = [(0, 0, 0), (0, 0, 255), (165, 42, 42), (127, 0, 255),
              (255, 255, 0), (231, 84, 128)]  # black,blue,brown,purple,pink,yellow
SOURCE_COLOR = (231, 84, 128)  # pink
DEST_COLOR = (255, 255, 0)  # yellow
SCREEN_SET_BACK = 5
CONFIG_LAYOUT_GRID_PATH = 0
CONFIG_LAYOUT_PATH = 1
CONFIG_GRID_PATH = 0


def DatabaseInput():
    #database inputs
    print("Creating database................Please wait.")
    df_comp = pd.read_csv(COMP_PATH)
    df_part = pd.read_csv(PARTLIST_PATH)
    input.Creatsamplefunction(df_comp,df_part,WORKCENTRE)

    # read samplefunction
    print("Creating Sample function................Please wait.")
    df1 = pd.read_csv(SAMPLEFUNCTION_PATH)
    input.CreateInputSheet(df1,SOURCE_TEXT_ARRAY,WORKCENTRE)
        


# To store coordinates of excel as point objects
class Point:
    def __init__(self, x: int, y: int):
        self.x = x
        self.y = y

# A data structure for queue used in bfs


class QueueNode:
    def __init__(self, point, dist, G_score, path, dest_pt):
        self.pt = point  # The point of the cell
        self.dist = dist  # Cell's distance from the source
        self.G_score = G_score  # cost needed to reach that path
        self.end_path = path  # Travelled path
        #self.end_dist= ((point.x-dest_pt.x)**2+(point.y-dest_pt.y)**2)**0.5
        self.x = abs(point.x-dest_pt.x)
        self.y = abs(point.y-dest_pt.y)
        # Cell's distance from the source
        self.end_dist = self.x+self.y+(1.414-2)*min(self.x, self.y)

        self.total = self.G_score+self.end_dist  # Travelled path

# Check whether given cell(row,col)
# is a valid cell or not


def isValid(row: int, col: int):
    return (row >= 0) and (row < ROW) and (col >= 0) and (col < COL)


# check whether given cell(row_,col_) have a gap of CHECKCELL
def checkWidth(row_, col_, k, dest_pt, grid):
    flag = 1

    # dictionary to remove indices from rowNum to not check backward cells
    index_list = {0: 3, 1: 2, 2: 1, 3: 0, 4: 7, 5: 6, 6: 5, 7: 4}
    diagonal_list = {0: (4, 5), 1: (4, 6), 2: (        # dictionary to remove indices from rowNum to not check diagonally backward cells
        5, 7), 3: (6, 7)}

    rowNum = [-1, 0, 0, 1, 1, 1, -1, -1]
    colNum = [0, -1, 1, 0, 1, -1, 1, -1]
    rowNum = numpy.array(rowNum, int)
    colNum = numpy.array(colNum, int)

    # delete diagonal elements from rownum and colnumn from diagonal_list
    if k < 4:
        rowNum = numpy.delete(rowNum, diagonal_list[k])
        colNum = numpy.delete(colNum, diagonal_list[k])
    # delete transverse,longitudnal elements from rownum and colnumn in index_list
    rowNum = numpy.delete(rowNum, index_list[k])
    colNum = numpy.delete(colNum, index_list[k])

    conversion_factor = 1
    for i in range(1, CHECK_CELL+1):
        # increase the checking range by i
        rowNum = (rowNum/conversion_factor)*i
        rowNum = rowNum.astype(int)
        colNum = (colNum/conversion_factor)*i
        colNum = colNum.astype(int)
        conversion_factor = i
        j = 0

        while (j < len(rowNum)):
            pt_x = row_+rowNum[j]
            pt_y = col_+colNum[j]
            if isValid(pt_x, pt_y):
                if grid[pt_x, pt_y].fill.fgColor.rgb in BLOCK_COLOR:
                    flag = 0  # set flag=0 when cells are blocked
                    break
                if grid[pt_x, pt_y].fill.fgColor.rgb in BIN_COLOR[:, 0]:
                    if pt_x == dest_pt.x and pt_y == dest_pt.y:
                        # remove transverse elements if destination is reached
                        rowNum = numpy.delete(rowNum, j)
                        colNum = numpy.delete(colNum, j)
                        # remove diagonal element if destination is reached
                        if k <= 3:
                            rowNum = numpy.delete(rowNum, (2, 3))
                            colNum = numpy.delete(colNum, (2, 3))
                        else:
                            if k >= 6:
                                if j == 0:
                                    rowNum = numpy.delete(rowNum, (4, 5))
                                    colNum = numpy.delete(colNum, (4, 5))
                                elif j == 1:
                                    rowNum = numpy.delete(rowNum, (3, 5))
                                    colNum = numpy.delete(colNum, (3, 5))
                                elif j == 2:
                                    rowNum = numpy.delete(rowNum, (3, 4))
                                    colNum = numpy.delete(colNum, (3, 4))

                            else:
                                if j == 3:
                                    rowNum = numpy.delete(rowNum, (3, 4))
                                    colNum = numpy.delete(colNum, (3, 4))
                                elif j == 1:
                                    rowNum = numpy.delete(rowNum, (4, 5))
                                    colNum = numpy.delete(colNum, (4, 5))
                                elif j == 2:
                                    rowNum = numpy.delete(rowNum, (3, 5))
                                    colNum = numpy.delete(colNum, (3, 5))

                        j = j-1
                        flag = 1
                    else:
                        flag = 0
                        break
            j = j+1
        if flag == 0:
            break

    return flag


# for multithreading inside main logic and best path
def globalParameter(row, col, block_color, bin_color):
    global ROW, COL, BLOCK_COLOR, BIN_COLOR
    ROW = row
    COL = col
    BLOCK_COLOR = block_color
    BIN_COLOR = bin_color


# for multi-Threading
def argumentGenerator(args):
    return breadthFirstSearch(*args)


# Function to find the shortest path_script between
# a given source cell to a destination cell.
def breadthFirstSearch(src, dest, cells):
    rowNum = [-1, 0, 0, 1, 1, 1, -1, -1]
    colNum = [0, -1, 1, 0, 1, -1, 1, -1]
    # build matrix of rows and columns

    # Create a queue for bfs
    q = []

    # Distance of source cell is 0
    s = QueueNode(src, 0, 0, [src], dest)
    # Create an Open list and close list
    open_list = {}
    open_list[str(src.x)+str(src.y)] = s
    close_list = {}
    # Do a bfs starting from source cell
    while len(open_list):
        q = open_list.values()
        q = list(q)
        q = numpy.array(q)
        distance = [element.total for element in q]
        distance = numpy.array(distance)
        q = q[distance.argsort()]  # arranging them in descending order
        curr = q[0]  # Dequeue the front cell

        pt = curr.pt
        close_list[str(pt.x)+str(pt.y)] = curr

        # q=numpy.delete(q,0)
        del open_list[str(pt.x)+str(pt.y)]
        # If we have reached the destination cell,
        # we are done
        if pt.x == dest.x and pt.y == dest.y:
            return curr.dist, curr.end_path

        # Otherwise enqueue its adjacent cells
        for i in range(8):
            row = pt.x + rowNum[i]
            col = pt.y + colNum[i]

            # if adjacent cell is valid, has path_script
            # and not visited yet, enqueue it.
            # check whether the path is covered before to avoid backtracking
            if str(row)+str(col) not in close_list:
                if ((isValid(row, col) and not(cells[row, col].fill.fgColor.rgb in BLOCK_COLOR) and
                    not(cells[row, col].fill.fgColor.rgb in BIN_COLOR[:, 0]) and checkWidth(row, col, i, dest, cells))
                        or (row == dest.x and col == dest.y)):

                    if i < 4:  # non diagonal movements
                        G_score = 1

                        Adjcell = QueueNode(Point(row, col),
                                            curr.dist+1, curr.G_score+G_score, curr.end_path + [Point(row, col)], dest)

                        if str(row)+str(col) not in open_list:
                            open_list[str(row)+str(col)] = Adjcell
                        else:
                            if Adjcell.G_score < open_list[str(row)+str(col)].G_score:
                                open_list[str(row)+str(col)] = Adjcell
                    else:  # diagonal movements
                        G_score = 1.414  # diagonal distance

                        Adjcell = QueueNode(Point(row, col),
                                            curr.dist+1.414, curr.G_score+G_score, curr.end_path + [Point(row, col)], dest)

                        if str(row)+str(col) not in open_list:
                            open_list[str(row)+str(col)] = Adjcell
                        else:
                            # to check if the pt exist before nd if its check distance is lower or not
                            if Adjcell.G_score < open_list[str(row)+str(col)].G_score:
                                open_list[str(row)+str(col)] = Adjcell

    # Return -1 if destination cannot be reached
    return -1, curr.end_path


# Layout Reading
def readLayoutFromExcel(path, START_CELL_, END_CELL_):
    book = load_workbook(path.lstrip("\u202a"))
    print("Fetching layout from LTAHighrise-Trackout sheet.......")
    sheet = book["LTAHighrise-Trackout"]
    cells = sheet[START_CELL_ + ":" + END_CELL_]
    cells = numpy.array(cells)
    return cells, book


# for report1
def globalVar(BLOCK_COLOR_, BIN_COLOR_, ROW_, COL_):
    global BLOCK_COLOR, BIN_COLOR, ROW, COL
    BLOCK_COLOR = BLOCK_COLOR_
    BIN_COLOR = BIN_COLOR_
    ROW = ROW_
    COL = COL_


# findcolors and bin locations from  layout report
def findColor(book_):
    print("Fetching address of source , destination , block_colors from Grid_Loc_Highrise sheet.......")
    sheet_loc = book_[SHEET_NAME]
    max_row = sheet_loc.max_row
    bin_color = []

    block_color = []
    for src_end in range(1, max_row+1):
        if sheet_loc.cell(src_end, 1).value == None:
            break
    bin_array = "a3:b" + str(src_end-1)  # bin location starting from a3

    for dest_end in range(1, max_row+1):
        if sheet_loc.cell(dest_end, 4).value == None:
            break

    source_array = "d1:e" + str(dest_end-1)  # source loations starting From d1

    search_location = [sheet_loc[bin_array], sheet_loc[source_array]]

    # find bin and block color
    color_array = "a" + str(src_end) + ":c" + str(max_row)
    color_cells = sheet_loc[color_array]
    #color_cells = sheet_loc["a60:c69"]
    color_cells = numpy.array(color_cells)
    for row in color_cells:
        if(row[1].value == "Blocked"):
            block_color = block_color + [row[0].fill.fgColor.rgb]
        elif(not row[1].value == None):
            bin_color = bin_color + \
                [[row[0].fill.fgColor.rgb, row[1].value, row[2].value]]
    bin_color = numpy.array(bin_color)

    return search_location, block_color, bin_color


# Check whethere input locations are avaialble in layout
def findVariable(search_array, SOURCE_TEXT_, DEST_TEXT_):
    source = []
    dest = []
    # find source
    for values in search_array:
        flag = 0
        for row in values:
            if (row[0].value == SOURCE_TEXT_):
                source = row[1].value
                flag = 1
                break
        if flag:
            break
    if not flag:
        print(f"{SOURCE_TEXT} couldnt be found in GRID_LOC_HIGHRISE sheet")

    # find destination
    for values in search_array:
        flag = 0
        j = 0

        while j < len(DEST_TEXT_):
            for row in values:
                if (row[0].value == DEST_TEXT_[j]):
                    dest = dest + [[row[1].value, row[0].fill.fgColor.rgb]]
                    flag = 1
                    DEST_TEXT_.pop(j)
                    break
            if not flag:
                j = j+1
    if len(DEST_TEXT_):
        print(f"{DEST_TEXT_} couldnt be found in GRID_LOC_HIGHRISE sheet")

    return source, dest


def findColumnNumber(address):
    return column_index_from_string(address)

# convert to relative coordiantes respect to start and end


def findRelativeCoordinates(source_, dest_, START_CELL_):
    source_y, source_x = re.findall(
        '([a-z]+)([0-9]+)', source_, re.IGNORECASE)[0]
    dest_y, dest_x = re.findall('([a-z]+)([0-9]+)', dest_, re.IGNORECASE)[0]
    start_cell_y, start_cell_x = re.findall(
        '([a-z]+)([0-9]+)', START_CELL_, re.IGNORECASE)[0]

    # convert columns to numbers
    source_y = findColumnNumber(source_y)
    dest_y = findColumnNumber(dest_y)
    start_cell_y = findColumnNumber(start_cell_y)

    # relative coordinates with respect to start_range
    source_rel_x = int(source_x)-int(start_cell_x)
    source_rel_y = int(source_y)-int(start_cell_y)
    dest_rel_x = int(dest_x)-int(start_cell_x)
    dest_rel_y = int(dest_y)-int(start_cell_y)
    return source_rel_x, source_rel_y, dest_rel_x, dest_rel_y


# check valid source and destination
def checkSourceDest(src_pt_, dest_pt_, cells):
    if not (isValid(src_pt_.x, src_pt_.y)):
        return -1
    if not (isValid(dest_pt_.x, dest_pt_.y)):
        return -2
    rowNum = [-1, 0, 0, 1]
    colNum = [0, -1, 1, 0]

    # source check
    for i in range(4):
        row = src_pt_.x + rowNum[i]
        col = src_pt_.y + colNum[i]
        flag = 1
        if((isValid(row, col) and not(cells[row, col].fill.fgColor.rgb in BLOCK_COLOR) and
                not(cells[row, col].fill.fgColor.rgb in BIN_COLOR[:, 0]) and checkWidth(row, col, i, dest_pt_, cells))
                or (row == dest_pt_.x and col == dest_pt_.y)):
            flag = 0
            break
    if(flag):
        return -3
    # destination check
    for i in range(4):
        row = dest_pt_.x + rowNum[i]
        col = dest_pt_.y + colNum[i]
        flag = 1
        if(isValid(row, col) and not(cells[row, col].fill.fgColor.rgb in BLOCK_COLOR) and
           not(cells[row, col].fill.fgColor.rgb in BIN_COLOR[:, 0]) and checkWidth(row, col, i, dest_pt_, cells)):
            flag = 0
            break
    if(flag):
        return -4
    return 1


# input workbook sample
def readReport2excel(path, mech_name, source_text):

    # return all_path
    dest_array = []
    book = load_workbook(path.lstrip("\u202a"))
    print("Fetching destination locations from Report Optimization workbook.......")
    sheet = book["Sheet1"]
    max_row = sheet.max_row
    cells = sheet["a1:F"+str(max_row)]
    flag = 0  # to check if station any station is there or not

    #ILC
    book_ILC = load_workbook(EXCEL_PATH.lstrip("\u202a"))
    sheet_loc_ILC = book_ILC[SHEET_NAME]
    max_row_ILC = sheet_loc_ILC.max_row
    for src_end in range(1, max_row_ILC+1):
        if sheet_loc_ILC.cell(src_end, 1).value == None:
            break
    bin_array = "a3:b" + str(src_end-1)
    

    for row in cells:

        if row[0].value == source_text and row[1].value == mech_name:
            if ICL_check(str(row[5].value),sheet_loc_ILC[bin_array]):
                dest_array.append([str(row[5].value), str(row[3].value)])
                mech_class_name = row[1].value
                serial_number = row[2].value
                flag = 1
    if flag:
        return dest_array, mech_class_name, serial_number
    return -1, -1, -1


class BestPathParamters():
    def __init__(self, x, y, dist, end_path):
        self.x = x
        self.y = y
        self.dist = dist  # distance from 1st pt to another
        self.end_path = end_path  # path taken


# check if closed path formed or not, returns 1 if closed path exist otherwise 0
def CircularCheck(x, y, path):
    check_path = path
    global_y = y
    check_element_x = x

    while len(check_path):
        flag_x = 0
        for count, elements in enumerate(check_path):
            if check_element_x in [elements.x, elements.y]:
                if check_element_x == elements.x:
                    check_element_x = elements.y
                else:
                    check_element_x = elements.x
                check_path.pop(count)
                flag_x = 1
                break

        if check_element_x == global_y and flag_x == 1:
            return 1
        elif flag_x == 1:
            pass
        elif flag_x == 0:
            return 0
    return 0


# determine which location to pick first
def bestPathThread(dest_list_rel, grid, not_reached):
    print("Finding best path.......")

    visited = {}
    matrix = []
    dictionary = {}

    for count, i in enumerate(dest_list_rel):
        dictionary[i] = count+1

    a = [(i, j, grid, dictionary[i], dictionary[j])
         for count, i in enumerate(dest_list_rel) for j in dest_list_rel[count+1:]]
    a = numpy.array(a, dtype=object)
    with Pool(4, initializer=globalParameter, initargs=(ROW, COL, BLOCK_COLOR, BIN_COLOR)) as pool3:

        result3 = pool3.map(argumentGenerator, a[:, 0:3])

    result3 = list(result3)
    result3 = numpy.array(result3, dtype=object)
    results3 = numpy.append(result3, a[:, 3:], axis=1)
    for value in results3:
        if value[0] > -1:
            p = BestPathParamters(value[2], value[3], value[0], value[1])
            matrix.append(p)
            visited[value[2]] = 1
            visited[value[3]] = 1

    for i in range(1, len(dest_list_rel)+1):
        if i not in visited.keys():
            print(
                f"{dest_name_ALL[i-1]}-{dest_list_address_ALL[i-1][0]} cant be reached")
            not_reached = not_reached + \
                [[dest_list_rel[i-1], dest_name_ALL[i-1]]]

    # sort the distance in ascending order
    sort_matrix = [element.dist for element in matrix]
    sort_matrix = numpy.array(sort_matrix)
    matrix = numpy.array(matrix)
    matrix = matrix[sort_matrix.argsort()]

    best_path = []
    check_list = []  # to check if the add path is repeated again
    best_path.append(matrix[0])
    check_list = check_list+[matrix[0].x]
    check_list = check_list+[matrix[0].y]
    matrix = numpy.delete(matrix, 0)

    # find best_path algorithm
    while len(best_path) < len(dest_list_rel)-1:
        flag = 1
        if (((Counter(check_list)[matrix[0].x] > 1) or (Counter(check_list)[matrix[0].y] > 1)) or CircularCheck(matrix[0].x, matrix[0].y, copy.copy(best_path))):
            matrix = numpy.delete(matrix, 0)
        else:
            check_list = check_list+[matrix[0].x]
            check_list = check_list+[matrix[0].y]
            best_path = best_path+[matrix[0]]
            matrix = numpy.delete(matrix, 0)

        if not len(matrix):
            return best_path, not_reached

    return best_path, not_reached


# arranging the best paths
def bestPathOrder(best_path):
    print("Finding appropriate order of arrangment for best path.......", "\n")
    oldbestpath = best_path
    # append first element from old best path to new one
    newbestpath = [oldbestpath[0]]
    oldbestpath.pop(0)
    search_element = newbestpath[0]  # make this as search element
    insert = 0
    loop = 1
    # loop to search the search element in old best path
    while len(oldbestpath):
        flag = 0
        for count, elements in enumerate(oldbestpath):
            if search_element.x in [elements.x, elements.y] or search_element.y in [elements.x, elements.y]:
                # (1-3,2-3) we should change them as (1-3.3-2)
                if search_element.x == elements.x or search_element.y == elements.y:
                    if insert:
                        elements.end_path = elements.end_path[::-1]
                        swap = elements.x
                        elements.x = elements.y
                        elements.y = swap
                        newbestpath.insert(0, elements)
                    else:
                        elements.end_path = elements.end_path[::-1]
                        swap = elements.x
                        elements.x = elements.y
                        elements.y = swap
                        newbestpath = newbestpath+[elements]
                else:  # (1-2,2-3)
                    if insert:
                        newbestpath.insert(0, elements)
                    else:
                        newbestpath = newbestpath+[elements]
                oldbestpath.pop(count)
                flag = 1
                break
        # if no seaarch element is found make this as the last path
        if flag == 0 and len(oldbestpath):
            search_element = newbestpath[0]
            insert = 1

        elif len(oldbestpath):
            if insert:
                search_element = newbestpath[0]
            else:
                search_element = newbestpath[-1]

    return newbestpath


class MechClass():
    def __init__(self, mech_class_name, station_name, shortest_dist_All, shortest_dist_ONE, end_path_array_all, end_path_array_one, not_reached, data_one, data_all, item_map, serial_number, item_dict, material_picked, not_available):
        self.station_name = station_name
        self.mech_class_name = mech_class_name
        self.distance_all = shortest_dist_All
        self.distance_one = shortest_dist_ONE
        self. end_path_array_all = end_path_array_all
        self.end_path_array_one = end_path_array_one
        self.not_reached = not_reached
        self.data_one = data_one
        self.data_all = data_all
        self.item_map = item_map
        self.serial_number = serial_number
        self.material_available = item_dict
        self.material_picked = material_picked
        self.not_available = not_available

# for multithreading inside main fn


def globalExcel(EXCEL_PATH, mech_name):
    global grid, book, ROW, COL, MECH_NAME
    grid, book = readLayoutFromExcel(
        EXCEL_PATH, START_CELL, END_CELL)
    ROW = grid.shape[0]  # total rows
    COL = grid.shape[1]
    MECH_NAME = mech_name


def main():
    mech_all = []
    for MECH_NAME in MECH_CLASSES:
        mech = []
        globalExcel(EXCEL_PATH, MECH_NAME)
        with Pool(3, initializer=globalExcel, initargs=(EXCEL_PATH, MECH_NAME)) as pool:

            results = pool.map(mainLogic, SOURCE_TEXT_ARRAY)
        for result in results:
            if result != -1:
                mech = mech+[result]
        mech_all = mech_all+[mech]
    return mech_all, grid

def ICL_check(dest_list_name_ilc_,ILC_array):

    for row in ILC_array:
        if str(row[0].value)==dest_list_name_ilc_ and row[0].fill.fgColor.rgb=="FF0070C0":
            return True

    return False




def mainLogic(SOURCE):
    global ROW, COL, BLOCK_COLOR, BIN_COLOR, dest_name_ALL, dest_list_address_ALL, SOURCE_TEXT, SEARCH_ARRAY

    not_reached = []
    SOURCE_TEXT = SOURCE
    print(f"Evaluating for {SOURCE_TEXT} station")

    dest_list_name, mech_name, serial_number = readReport2excel(
        REPORT_PATH, MECH_NAME, SOURCE_TEXT)  # (ssc00123)

    if dest_list_name == -1:
        return -1
    SEARCH_ARRAY, BLOCK_COLOR, BIN_COLOR = findColor(book)
    item_dict = {}  # stores materials as key
    for rows in dest_list_name:
        if rows[1] in item_dict:
            item_dict[str(rows[1])] = item_dict[str(rows[1])]+[rows[0]]
        else:
            item_dict[str(rows[1])] = [rows[0]]
    material_dict = {}  # store bin locations as key
    material_picked = {}
    not_available = []
    for rows in item_dict:
        maximum_distance = []
        a = []
        if len(item_dict[rows]) > 1:
            flag = 0

            for address in item_dict[rows]:
                source, dest_list_address = findVariable(SEARCH_ARRAY,
                                                         SOURCE_TEXT, [address])

                if len(dest_list_address) == 0:
                    continue

                source_rel_x, source_rel_y, dest_rel_x, dest_rel_y = findRelativeCoordinates(
                    source, dest_list_address[0][0], START_CELL)

                src_pt = Point(source_rel_x, source_rel_y)
                dest_pt = Point(dest_rel_x, dest_rel_y)

                check = checkSourceDest(src_pt, dest_pt, grid)
                if check == -2:
                    continue
                elif check == -4:
                    continue
                flag = 1
                a = a+[[src_pt, dest_pt, grid, address, rows]]
                # dist,end_path=breadthFirstSearch(src_pt,dest_pt,grid)
                # maximum_distance.append([dist,address,rows])
            a = numpy.array(a, dtype=object)

            if len(a):
                with Pool(4, initializer=globalParameter, initargs=(ROW, COL, BLOCK_COLOR, BIN_COLOR)) as pool1:

                    result1 = pool1.map(argumentGenerator, a[:, 0:3])

                results1 = list(result1)
                results1 = numpy.array(results1, dtype=object)

                maximum_distance = numpy.append(
                    results1[:, 0].reshape(len(results1), 1), a[:, 3:], axis=1)
                # maximum_distance=numpy.array(maximum_distance)
                # maximum_distance=maximum_distance[maximum_distance[:,0].astype('float').argsort()[::-1]]
                maximum_distance = maximum_distance[maximum_distance[:, 0].astype(
                    'float').argsort()]  # minimum distance of all items to be picked

                first_row = -1
                # to ignore distance valuewhich is -1
                for i in range(0, len(maximum_distance)):
                    if not maximum_distance[i, 0] == -1:
                        first_row = i
                        break

                # all the distance returns -1 add to the material dict
                if first_row == -1:
                    for address2 in item_dict[rows]:
                        if address1 in material_dict:
                            material_dict[address2] = material_dict[address2]+[rows]
                        else:
                            material_dict[address2] = [rows]
                    continue

                if maximum_distance[first_row, 1] in material_dict:
                    material_dict[maximum_distance[first_row, 1]
                                  ] = material_dict[maximum_distance[first_row, 1]] + [maximum_distance[first_row, 2]]
                else:
                    material_dict[maximum_distance[first_row, 1]] = [
                        maximum_distance[first_row, 2]]
                material_picked[rows] = maximum_distance[first_row, 1]

                # if maximum_distance[0,1] in material_dict:
                #material_dict[maximum_distance[0,1]]=material_dict[maximum_distance[0,1]] + [maximum_distance[0,2]]
               # else:
                # material_dict[maximum_distance[0,1]]=[maximum_distance[0,2]]
                # material_picked[rows]=maximum_distance[0,1]
            else:
                for address1 in item_dict[rows]:
                    source, dest_list_address = findVariable(SEARCH_ARRAY,
                                                             SOURCE_TEXT, [address1])

                    if len(dest_list_address) == 0:
                        not_available = not_available + [[rows, address1]]
                        continue
                    if address1 in material_dict:
                        material_dict[address1] = material_dict[address1]+[rows]
                    else:
                        material_dict[address1] = [rows]

        else:
            source, dest_list_address = findVariable(SEARCH_ARRAY,
                                                     SOURCE_TEXT, [item_dict[rows][0]])

            if len(dest_list_address) == 0:
                not_available = not_available + [[rows, item_dict[rows][0]]]
                continue

            if item_dict[rows][0] in material_dict:
                material_dict[item_dict[rows][0]
                              ] = material_dict[item_dict[rows][0]] + [rows]
            else:
                material_dict[item_dict[rows][0]] = [rows]
            material_picked[rows] = item_dict[rows][0]

    dest_list_name = list(material_dict.keys())

    # find source address ,block colors and destination address
    source, dest_list_address = findVariable(SEARCH_ARRAY,
                                             SOURCE_TEXT, copy.copy(dest_list_name))

    dest_list_dict = {}
    dest_name = {}
    dest_list_dict["ALL"] = []
    dest_name["ALL"] = []
    dest_list_dict["ONE"] = []
    dest_name["ONE"] = []

    for count, element in enumerate(dest_list_address):
        for row in BIN_COLOR:
            if row[0] == element[1]:
                if row[2] == "ALL":
                    dest_list_dict[row[2]] = dest_list_dict[row[2]] + [element]
                    dest_name[row[2]] = dest_name[row[2]] + \
                        [dest_list_name[count]]
                else:
                    dest_list_dict[row[2]] = dest_list_dict[row[2]] + \
                        [element]*len(material_dict[dest_list_name[count]])
                    dest_name[row[2]] = dest_name[row[2]] + [dest_list_name[count]
                                                             ]*len(material_dict[dest_list_name[count]])
                break

    dist_array_all = []  # distance betweena all points
    end_path_array_all = []  # travel path between all points
    dest_list_rel = []  # destination points relative to the startt_range
    # path that couldnt be reached

    dest_name_ALL = copy.copy(dest_name["ALL"])
    dest_name_ONE = copy.copy(dest_name["ONE"])
    dest_list_address_ALL = copy.copy(dest_list_dict["ALL"])
    dest_list_address_ONE = copy.copy(dest_list_dict["ONE"])
    data_all = []
    shortest_dist_All = []
    if len(dest_list_address_ALL):
        count = 0
        while count < len(dest_list_address_ALL):
            # find source,destination coordinates
            source_rel_x, source_rel_y, dest_rel_x, dest_rel_y = findRelativeCoordinates(
                source, dest_list_address_ALL[count][0], START_CELL)

        # point object
            src_pt = Point(source_rel_x, source_rel_y)
            dest_pt = Point(dest_rel_x, dest_rel_y)

        # check valid source and destination
            check = checkSourceDest(src_pt, dest_pt, grid)
            if check == -1:
                print(f"{SOURCE_TEXT}-{source} is out of range")
                exit()
            elif check == -3:
                print(f"{SOURCE_TEXT}-{source} cant be reached")
                exit()
            elif check == -2:
                print(
                    f"{dest_name_ALL[count]}-{dest_list_address_ALL[count][0]} is out of range")
                not_reached = not_reached+[[dest_pt, dest_name_ALL[count]]]
                dest_name_ALL.pop(count)
                dest_list_address_ALL.pop(count)
                continue
            elif check == -4:
                print(
                    f"{dest_name_ALL[count]}-{dest_list_address_ALL[count][0]} cant be reached")
                not_reached = not_reached+[[dest_pt, dest_name_ALL[count]]]
                dest_name_ALL.pop(count)
                dest_list_address_ALL.pop(count)
                continue
            dest_list_rel = dest_list_rel + [dest_pt]
            count = count+1

        # find best path
        if len(dest_list_rel) > 1:
            best_path, not_reached = bestPathThread(
                copy.copy(dest_list_rel), grid, not_reached)
            # find best path with correct order
            newbestpath = bestPathOrder(copy.copy(best_path))

            if len(dest_list_rel) > 2:
                if newbestpath[0].y == newbestpath[1].x:  # (1-2,2_4,3-4)
                    pass
                else:
                    # (3-4,2-4,1-2) then reverse the order as (1-2,2_4,3-4)
                    newbestpath = newbestpath[::-1]

            dest_name_ALL_copy = []
            for count, element in enumerate(newbestpath):
                end_path_array_all = end_path_array_all+[element.end_path]
                dist_array_all = dist_array_all+[element.dist]
                dest_name_ALL_copy = dest_name_ALL_copy + \
                    [dest_name_ALL[element.x-1]]
                if count == len(newbestpath)-1:
                    dest_name_ALL_copy = dest_name_ALL_copy + \
                        [dest_name_ALL[element.y-1]]

            # find the distnace between aources and newbestpath[0] and newbestpath[-1]
            dest_pt_1 = newbestpath[0].end_path[0]
            dest_pt_2 = newbestpath[-1].end_path[-1]
            dist1, end_path1 = breadthFirstSearch(dest_pt_1, src_pt, grid)
            dist2, end_path2 = breadthFirstSearch(dest_pt_2, src_pt, grid)

            dest_name_ALL = dest_name_ALL_copy
            if dist1 <= dist2:
                dest_name_ALL.insert(0, SOURCE_TEXT)
                dest_name_ALL.append(SOURCE_TEXT)

                end_path1 = end_path1[::-1]
                end_path_array_all.insert(0, end_path1)
                dist_array_all.insert(0, dist1)
                end_path_array_all = end_path_array_all+[end_path2]
                dist_array_all = dist_array_all + [dist2]

            else:
                dest_name_ALL = dest_name_ALL[::-1]
                dest_name_ALL.insert(0, SOURCE_TEXT)
                dest_name_ALL.append(SOURCE_TEXT)

                end_path_array_all = end_path_array_all[::-1]
                dist_array_all = dist_array_all[::-1]

                for count, element in enumerate(end_path_array_all):
                    end_path_array_all[count] = end_path_array_all[count][::-1]
                end_path2 = end_path2[::-1]
                end_path_array_all.insert(0, end_path2)
                dist_array_all.insert(0, dist2)
                end_path_array_all = end_path_array_all+[end_path1]
                dist_array_all = dist_array_all + [dist1]

            for count, value in enumerate(dist_array_all):
                data_all = data_all + \
                    [[dest_name_ALL[count], dest_name_ALL[count+1],
                        round(value*SCALE), round(value*SCALE/STRIDE)]]

        elif len(dest_list_rel) > 0:
            dest_pt_1 = dest_list_rel[0]
            dist1, end_path1 = breadthFirstSearch(src_pt, dest_pt_1, grid)
            end_path_array_all = end_path_array_all+[end_path1]
            dist_array_all = dist_array_all + [dist1]
            data_all = data_all + [[SOURCE_TEXT, dest_name_ALL[0], round(
                dist_array_all[-1]*SCALE), round(dist_array_all[-1]*SCALE/STRIDE)]]
            end_path_array_all = end_path_array_all+[end_path1[::-1]]
            dist_array_all = dist_array_all + [dist1]
            data_all = data_all + [[dest_name_ALL[0], SOURCE_TEXT, round(
                dist_array_all[-1]*SCALE), round(dist_array_all[-1]*SCALE/STRIDE)]]

        shortest_dist_All = list(map(lambda x: round(x*SCALE), dist_array_all))

    dist_array_one = []  # distance betweena all points
    end_path_array_one = []
    data_one = []
    shortest_dist_ONE = []
    dummy_data = []

    if len(dest_list_address_ONE):
        count = 0
        while count < len(dest_list_address_ONE):
            # find source,destination coordinates
            source_rel_x, source_rel_y, dest_rel_x, dest_rel_y = findRelativeCoordinates(
                source, dest_list_address_ONE[count][0], START_CELL)

            # point object
            src_pt = Point(source_rel_x, source_rel_y)
            dest_pt = Point(dest_rel_x, dest_rel_y)

            # check valid source and destination
            check = checkSourceDest(src_pt, dest_pt, grid)
            if check == -1:
                print(f"{SOURCE_TEXT}-{source} is out of range")
                exit()
            elif check == -3:
                print(f"{SOURCE_TEXT}-{source} cant be reached")
                exit()
            elif check == -2:
                print(
                    f"{dest_name_ONE[count]}-{dest_list_address_ONE[count][0]} is out of range")
                not_reached = not_reached+[[dest_pt, dest_name_ONE[count]]]
                dest_name_ONE.pop(count)
                dest_list_address_ONE.pop(count)
                continue
            elif check == -4:
                print(
                    f"{dest_name_ONE[count]}-{ dest_list_address_ONE[count][0]} cant be reached")
                not_reached = not_reached+[[dest_pt, dest_name_ONE[count]]]
                dest_name_ONE.pop(count)
                dest_list_address_ONE.pop(count)
                continue
            dummy_data = dummy_data+[(src_pt, dest_pt, grid)]
            count = count+1

    with Pool(4, initializer=globalParameter, initargs=(ROW, COL, BLOCK_COLOR, BIN_COLOR)) as pool2:

        result2 = pool2.map(argumentGenerator, dummy_data)

    results2 = list(result2)

    for count, values in enumerate(results2):
        if values[0] >= 0:
            dist_array_one = dist_array_one+[values[0]]
            end_path_array_one = end_path_array_one+[values[1]]
            data_one = data_one + [[SOURCE_TEXT, dest_name_ONE[count], round(
                dist_array_one[-1]*SCALE), round(dist_array_one[-1]*SCALE/STRIDE)]]
            dist_array_one = dist_array_one+[values[0]]
            end_path_array_one = end_path_array_one+[values[1][::-1]]
            data_one = data_one + [[dest_name_ONE[count], SOURCE_TEXT, round(
                dist_array_one[-1]*SCALE), round(dist_array_one[-1]*SCALE/STRIDE)]]

        else:
            not_reached = not_reached + \
                [[dummy_data[count][1], dest_name_ONE[count]]]
            print(
                f"No path found for {dest_name_ONE[count]}-{dest_list_address_ONE[count][0]}")

    shortest_dist_ONE = list(map(lambda x: round(x*SCALE), dist_array_one))

    return MechClass(mech_name, SOURCE_TEXT, shortest_dist_All, shortest_dist_ONE,  end_path_array_all, end_path_array_one, not_reached, data_one, data_all, material_dict, serial_number, item_dict, material_picked, not_available)


if __name__ == "__main__":
    DatabaseInput()
    if MECH_CLASSES=="ALL":
        df1=pd.read_excel(REPORT_PATH)
        MECH_CLASSES=list(df1["Mech Class"].unique())
    SOURCE_TEXT_ARRAY= list(map(lambda x:WORKCENTRE[0].replace("16WS","") + "_"+ x,SOURCE_TEXT_ARRAY))
    if REPORT2:
        mech_all, grid = main()
        sub_directory_names = SOURCE_TEXT_ARRAY

        for mech in MECH_CLASSES:
            # shutil.rmtree(f"{mech}",ignore_errors=True)
            for name in sub_directory_names:
                shutil.rmtree(f"Report2/{mech}/{name}", ignore_errors=True)
                os.makedirs(f"Report2/{mech}/{name}", exist_ok=True)

        datadump.writeExcel(mech_all, sys.modules[__name__],WORKCENTRE)
        print("time taken is", time.time()-start)
        print("Visualizing......", "\n")
        report2pygame.visualize(grid, mech_all,
                                ROW, COL, sys.modules[__name__])

        # Consolidating GIF
        GIF_directory = "Report2GIF"
        for WORK in WORKCENTRE:
            GIF_directory = GIF_directory + "_" + WORK.replace("16WS","")
        shutil.rmtree(f"Report2/{GIF_directory}", ignore_errors=True)
        os.makedirs(f"Report2/{GIF_directory}", exist_ok=True)

        for mech in MECH_CLASSES:
            shutil.rmtree(
                f"Report2/{GIF_directory}/{mech}", ignore_errors=True)
            os.makedirs(f"Report2/{GIF_directory}/{mech}", exist_ok=True)
            for name in sub_directory_names:
                for files in glob.glob(f"Report2/{mech}/{name}/*.gif"):
                    shutil.copy(files, f"Report2/{GIF_directory}/{mech}")

    else:
        for mech in MECH_CLASSES:
            shutil.rmtree(f"Report1/{mech}", ignore_errors=True)
            sub_directory_names = SOURCE_TEXT_ARRAY
            for name in sub_directory_names:
                os.makedirs(f"Report1/{mech}", exist_ok=True)

        mech_all, grid, ROW, COL = Report1_thread.report1Main(MECH_CLASSES,SOURCE_TEXT_ARRAY)
        report1datadump.writeExcel(mech_all,WORKCENTRE)

        print("Visualizing......", "\n")
        report1pygame.visualize(grid, mech_all,
                                ROW, COL, sys.modules[__name__])

        # consolidating GIF

        IMAGE_DIRECTORY = "Report1ALL"
        for WORK in WORKCENTRE:
            IMAGE_DIRECTORY = IMAGE_DIRECTORY + "_" + WORK.replace("16WS","")
        shutil.rmtree(f"Report1/{IMAGE_DIRECTORY}", ignore_errors=True)
        os.makedirs(f"Report1/{IMAGE_DIRECTORY}", exist_ok=True)

        for mech in MECH_CLASSES:
            shutil.rmtree(
                f"Report1/{IMAGE_DIRECTORY}/{mech}", ignore_errors=True)
            os.makedirs(f"Report1/{IMAGE_DIRECTORY}/{mech}", exist_ok=True)
            for files in glob.glob(f"Report1/{mech}/*.jpg"):
                shutil.copy(files, f"Report1/{IMAGE_DIRECTORY}/{mech}")
