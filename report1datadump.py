from openpyxl import load_workbook, Workbook
import numpy


def writeExcel(mech_all,WORKCENTRE):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Reached"
    name = ""
    wb.create_sheet("not_reached")
    wb.create_sheet("not_available")
    sheet2 = wb['not_reached']
    sheet3 = wb['not_available']
    sheet.cell(1, 1).value = "Station"
    sheet.cell(1, 2).value = "Mech_Class"
    sheet.cell(1, 3).value = "Destination"
    sheet.cell(1, 4).value = "Distance (ft)"

    sheet2.cell(1, 1).value = "Station"
    sheet2.cell(1, 2).value = "Mech_Class"
    sheet2.cell(1, 3).value = "not_reached"

    sheet3.cell(1, 1).value = "Mech_Class"
    sheet3.cell(1, 2).value = "Bin location"

    i = 1
    j = 1
    name = name+"_"+ WORKCENTRE[0].replace("16WS","")
    for mech_class in mech_all:
        i = i+1
        
        for station in mech_class:
            for row in station.data_reached:
                sheet.cell(i, j).value = station.station_name
                sheet.cell(i, j+1).value = station.mech_class_name
                sheet.cell(i, j+2).value = row[0]
                sheet.cell(i, j+3).value = row[1]
                i = i+1
    i = 1
    for mech_class in mech_all:
        i = i+1

        for station in mech_class:
            for row in station.data_not_reached:
                sheet2.cell(i, j).value = station.station_name
                sheet2.cell(i, j+1).value = station.mech_class_name
                sheet2.cell(i, j+2).value = row
                i = i+1

    i = 2

    for mech_class in mech_all:
        # i=i+1
        for station in mech_class:
            for row in station.not_available:
                sheet3.cell(i, j).value = station.mech_class_name
                sheet3.cell(i, j+1).value = row

                i = i+1

    wb.save(f"Report1/Report1_data{name}.xlsx")
