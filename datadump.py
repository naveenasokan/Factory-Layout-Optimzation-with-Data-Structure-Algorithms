from openpyxl import load_workbook, Workbook
import numpy


def GlobalVariables(module):
    global STRIDE
    STRIDE = module.STRIDE


def writeExcel(mech_all, module,WORKCENTRE):
    GlobalVariables(module)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Data"
    name = ""
    wb.create_sheet("Not_Picked")
    wb.create_sheet("Picked")
    wb.create_sheet("Not Available")
    sheet2 = wb["Not_Picked"]
    sheet3 = wb["Picked"]
    sheet4 = wb["Not Available"]
    i = 1
    j = 1
    sheet.cell(1, 1).value = "Station"
    sheet.cell(1, 2).value = "Mech_Class"
    #sheet.cell(1,3).value="Serial Number"
    sheet.cell(1, 3).value = "Source"
    sheet.cell(1, 4).value = "Destination"
    sheet.cell(1, 5).value = "Item picked"
    sheet.cell(1, 6).value = "Distance (ft)"
    sheet.cell(1, 7).value = "Time (sec)"
    sheet.cell(1, 8).value = "Time (min)"

    sheet2.cell(1, 1).value = "Station"
    sheet2.cell(1, 2).value = "Mech_Class"
    #sheet2.cell(1,3).value="Serial Number"
    sheet2.cell(1, 3).value = "Item Not Picked"
    sheet2.cell(1, 4).value = "Bin location"

    sheet3.cell(1, 1).value = "Station"
    sheet3.cell(1, 2).value = "Mech_Class"
    #sheet3.cell(1,3).value="Serial Number"
    sheet3.cell(1, 3).value = "Item"
    sheet3.cell(1, 4).value = "Available locations"
    sheet3.cell(1, 5).value = "Picked locations"

    sheet4.cell(1, 1).value = "Mech_Class"
    sheet4.cell(1, 2).value = "Item not Avalible"
    sheet4.cell(1, 3).value = "Bin location"

# data
    i = 2
    name = name+"_"+ WORKCENTRE[0].replace("16WS","")
    for mech_class in mech_all:
        for station in mech_class:
            data_all = numpy.array(station.data_all)
            data_one = numpy.array(station.data_one)
            # i=i+1
            for row in station.data_all:
                sheet.cell(i, j).value = station.station_name
                sheet.cell(i, j+1).value = station.mech_class_name
               # sheet.cell(i,j+2).value=station.serial_number
                sheet.cell(i, j+2).value = row[0]
                sheet.cell(i, j+3).value = row[1]
                sheet.cell(i, j+5).value = row[2]
                sheet.cell(i, j+6).value = row[3]
                if row[1] in station.item_map:
                    sheet.cell(
                        i, j+4).value = ",".join(station.item_map[row[1]])

                i = i+1
            for row in station.data_one:
                sheet.cell(i, j).value = station.station_name
                sheet.cell(i, j+1).value = station.mech_class_name
                # sheet.cell(i,j+2).value=station.serial_number
                sheet.cell(i, j+2).value = row[0]
                sheet.cell(i, j+3).value = row[1]
                sheet.cell(i, j+5).value = row[2]
                sheet.cell(i, j+6).value = row[3]
                if row[1] in station.item_map:
                    sheet.cell(i, j+4).value = station.item_map[row[1]][0]
                    station.item_map[row[1]] = numpy.array(
                        station.item_map[row[1]])
                    station.item_map[row[1]] = numpy.roll(
                        station.item_map[row[1]], -1)

                i = i+1
            sheet.cell(i, 1).value = station.station_name
            sheet.cell(i, 2).value = station.mech_class_name
            sheet.cell(i, 5).value = "Total"
            # sheet.cell(i,3).value=station.serial_number
            if len(data_all) and len(data_one):
                sheet.cell(i, 6).value = round(numpy.sum(data_all[:, 2].astype(
                    'float')) + numpy.sum(data_one[:, 2].astype('float')))
                sheet.cell(i, 7).value = round(numpy.sum(data_all[:, 3].astype(
                    'float')) + numpy.sum(data_one[:, 3].astype('float')))
                sheet.cell(i, 8).value = round(((numpy.sum(data_all[:, 2].astype(
                    'float')) + numpy.sum(data_one[:, 2].astype('float')))/STRIDE)/60, 1)
            elif len(data_all):
                sheet.cell(i, 6).value = round(
                    numpy.sum(data_all[:, 2].astype('float')))
                sheet.cell(i, 7).value = round(
                    numpy.sum(data_all[:, 3].astype('float')))
                sheet.cell(i, 8).value = round(
                    (numpy.sum(data_all[:, 2].astype('float'))/STRIDE)/60, 1)
            elif len(data_one):
                sheet.cell(i, 6).value = round(
                    numpy.sum(data_one[:, 2].astype('float')))
                sheet.cell(i, 7).value = round(
                    numpy.sum(data_one[:, 3].astype('float')))
                sheet.cell(i, 8).value = round(
                    (numpy.sum(data_one[:, 2].astype('float'))/STRIDE)/60, 1)

            i = i+1

    i = 2

#not reached
    for mech_class in mech_all:
        # i=i+1
        for station in mech_class:
            if len(station.not_reached):
                not_reached = numpy.array(station.not_reached)
                not_reached = not_reached[:, 1]
                not_reached = list(set(list(not_reached)))
                for value in not_reached:

                    for not_picked in station.item_map[value]:
                        sheet2.cell(i, j).value = station.station_name
                        sheet2.cell(i, j+1).value = station.mech_class_name
                        # sheet2.cell(i,j+2).value=station.serial_number
                        sheet2.cell(i, j+2).value = not_picked
                        sheet2.cell(
                            i, j+3).value = ",".join(station.material_available[not_picked])
                        i = i+1

    i = 2
    # picked items
    for mech_class in mech_all:
        # i=i+1
        for station in mech_class:
            not_picked = []
            not_available = []
            if len(station.not_available):
                not_available = numpy.array(station.not_available)
                not_available = not_available[:, 0]
            if len(station.not_reached):
                not_reached = numpy.array(station.not_reached)
                not_reached = not_reached[:, 1]
                not_reached = list(set(list(not_reached)))
                for value in not_reached:
                    not_picked = not_picked+station.item_map[value]
            for items in station.material_available:
                if items not in not_picked and items not in not_available:
                    sheet3.cell(i, j).value = station.station_name
                    sheet3.cell(i, j+1).value = station.mech_class_name
                   # sheet3.cell(i,j+2).value=station.serial_number
                    sheet3.cell(i, j+2).value = items
                    sheet3.cell(
                        i, j+3).value = ",".join(station.material_available[items])
                    sheet3.cell(i, j+4).value = station.material_picked[items]
                    i = i+1

    i = 2
    #not avaialble
    for mech_class in mech_all:
        # i=i+1
        for station in mech_class:
            for row in station.not_available:
                sheet4.cell(i, j).value = station.mech_class_name
                sheet4.cell(i, j+1).value = row[0]
                sheet4.cell(i, j+2).value = row[1]
                i = i+1

    wb.save(f"Report2/Report2_data{name}.xlsx")
